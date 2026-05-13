#!/usr/bin/env python3
import argparse
from itertools import combinations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.cluster.hierarchy import dendrogram, linkage

ANCESTRIES = []
LOCAL_PAIRS = []

CLASSIFICATION_LABELS = {
    "no_clear_signal": "No clear ancestry-specific pattern",
}

CLASSIFICATION_NOTES = {
    "no_clear_signal": "No ancestry reached the support threshold under the current QC rules.",
}

CLASSIFICATION_COLORS = {
    "no_clear_signal": "#B8BEC8",
}


def build_classification_metadata(ancestries: list[str]):
    labels = {}
    notes = {}
    colors = {}

    palette = [
        "#5DA5DA",
        "#FAA43A",
        "#60BD68",
        "#F17CB0",
        "#B2912F",
        "#B276B2",
        "#DECF3F",
    ]

    for idx, ancestry in enumerate(ancestries):
        key = f"{ancestry}_specific"
        labels[key] = f"{ancestry}-specific"
        notes[key] = f"Signal passes the QC rules in {ancestry} only."
        colors[key] = palette[idx % len(palette)]

    pair_colors = ["#3CAEA3", "#A1B83A", "#B27652", "#7E9C6F", "#9D7A68"]
    shared_index = 0
    for subset_size in range(2, len(ancestries)):
        for subset in combinations(ancestries, subset_size):
            key = f"{'_'.join(subset)}_shared"
            labels[key] = f"{' + '.join(subset)} shared"
            notes[key] = f"Signal is shared between {', '.join(subset)}."
            colors[key] = pair_colors[shared_index % len(pair_colors)]
            shared_index += 1

    labels["shared_all"] = f"Shared across all {len(ancestries)} ancestries"
    notes["shared_all"] = "Signal is present across all configured ancestries."
    colors["shared_all"] = "#C76BA2"

    labels["no_clear_signal"] = "No clear ancestry-specific pattern"
    notes["no_clear_signal"] = "No ancestry reached the support threshold under the current QC rules."
    colors["no_clear_signal"] = "#B8BEC8"

    return labels, notes, colors


def parse_args():
    parser = argparse.ArgumentParser(
        description="Build a locus-level report that classifies loci as ancestry-specific or shared."
    )
    parser.add_argument("--input", required=True, help="Input merged.maf_gwas_summary.tsv")
    parser.add_argument("--output-locus", required=True, help="Output locus-level report TSV")
    parser.add_argument("--output-snp", required=True, help="Output filtered SNP-level TSV used for the report")
    parser.add_argument("--output-excel", required=True, help="Styled Excel workbook for the locus report")
    parser.add_argument("--output-dendrogram", required=True, help="Colorful SVG dendrogram for loci")
    parser.add_argument("--cohort-maf-min", type=float, required=True, help="Minimum cohort MAF to keep a SNP")
    parser.add_argument("--ancestry-maf-min", type=float, required=True, help="Minimum ancestry MAF in at least one ancestry")
    parser.add_argument("--local-ancestry-min-logp", type=float, required=True, help="Minimum -log10(local ancestry p-value) required before a SNP is considered for ancestry assignment")
    parser.add_argument("--dosage-gwas-max-p", type=float, required=True, help="Maximum ancestry dosage GWAS p-value for assigning a SNP to one or more ancestries")
    parser.add_argument("--min-supporting-snps", type=int, required=True, help="Minimum number of supporting SNPs to call an ancestry active at a locus")
    parser.add_argument("--clustering-method", default="ward", help="Scipy linkage method for the dendrogram")
    parser.add_argument("--clustering-metric", default="euclidean", help="Distance metric for the dendrogram")
    return parser.parse_args()


def classify_active_ancestries(active):
    if not active:
        return "no_clear_signal"
    if len(active) == 1:
        return f"{active[0]}_specific"
    if len(active) == len(ANCESTRIES):
        return "shared_all"
    return f"{'_'.join(active)}_shared"


def best_ancestry(row):
    finite = {
        ancestry: row[f"p_dosage_{ancestry}"]
        for ancestry in ANCESTRIES
        if pd.notna(row.get(f"p_dosage_{ancestry}"))
    }
    if not finite:
        return ""
    return min(finite, key=finite.get)


def summarise_genes(series: pd.Series) -> str:
    genes = [str(value).strip() for value in series.fillna("") if str(value).strip()]
    unique = []
    seen = set()
    for gene in genes:
        if gene not in seen:
            unique.append(gene)
            seen.add(gene)
    return ";".join(unique)


def chromosome_sort_parts(series: pd.Series) -> tuple[pd.Series, pd.Series]:
    values = series.astype(str).str.replace("chr", "", regex=False)
    numeric = pd.to_numeric(values, errors="coerce")
    rank = numeric.fillna(10_000)
    return rank, values


def format_region(chrom, start, end) -> str:
    if pd.isna(start) or pd.isna(end):
        return f"chr{chrom}"
    return f"chr{chrom}:{int(start):,}-{int(end):,}"


def load_table(path: str) -> tuple[pd.DataFrame, list[str], list[str]]:
    df = pd.read_csv(
        path,
        sep="\t",
        dtype={"chr": str, "rsid": str, "locus_name": str, "gene": str, "ref": str, "alt": str},
    )
    required = [
        "chr", "pos", "rsid", "locus_name", "gene", "ref", "alt",
        "cohort_maf",
    ]
    missing = [column for column in required if column not in df.columns]
    if missing:
        raise ValueError(f"Input summary is missing required columns: {missing}")

    ancestry_maf_cols = [col for col in df.columns if col.endswith("_maf") and col != "cohort_maf"]
    ancestry_order = [col[:-4] for col in ancestry_maf_cols]
    if not ancestry_order:
        raise ValueError("Input summary has no ancestry-specific MAF columns (expected <ANCESTRY>_maf).")

    for ancestry in ancestry_order:
        required_col_set = [f"p_dosage_{ancestry}", f"OR_dosage_{ancestry}"]
        missing_cols = [col for col in required_col_set if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Input summary is missing required columns for ancestry {ancestry}: {missing_cols}")

    local_pairs = sorted(col.replace("p_local_ancestry_", "") for col in df.columns if col.startswith("p_local_ancestry_"))
    if not local_pairs:
        raise ValueError("Input summary has no local ancestry p-value columns (expected p_local_ancestry_<PAIR>).")
    for pair in local_pairs:
        col = f"OR_local_ancestry_{pair}"
        if col not in df.columns:
            raise ValueError(f"Input summary is missing required columns: ['{col}']")

    df = df.copy()
    df["chr"] = df["chr"].astype(str).str.replace("chr", "", regex=False)
    df["pos"] = pd.to_numeric(df["pos"], errors="coerce")

    numeric_columns = [column for column in required if column not in {"chr", "rsid", "locus_name", "gene", "ref", "alt"}]
    numeric_columns.extend(ancestry_maf_cols)
    numeric_columns.extend([f"p_dosage_{ancestry}" for ancestry in ancestry_order])
    numeric_columns.extend([f"OR_dosage_{ancestry}" for ancestry in ancestry_order])
    numeric_columns.extend([f"p_local_ancestry_{pair}" for pair in local_pairs])
    numeric_columns.extend([f"OR_local_ancestry_{pair}" for pair in local_pairs])
    for column in numeric_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    return df.dropna(subset=["chr", "pos", "locus_name", "rsid"]).copy(), ancestry_order, local_pairs


def add_filter_columns(df: pd.DataFrame, args) -> pd.DataFrame:
    df = df.copy()
    df["pass_cohort_maf"] = df["cohort_maf"] >= args.cohort_maf_min
    df["max_ancestry_maf"] = df[[f"{ancestry}_maf" for ancestry in ANCESTRIES]].max(axis=1, skipna=True)
    df["pass_ancestry_maf"] = df["max_ancestry_maf"] >= args.ancestry_maf_min

    for pair in LOCAL_PAIRS:
        local_p = pd.to_numeric(df[f"p_local_ancestry_{pair}"], errors="coerce").clip(lower=1e-300)
        df[f"local_logp_{pair}"] = -np.log10(local_p)

    local_logp_columns = [f"local_logp_{pair}" for pair in LOCAL_PAIRS]
    df["best_local_logp"] = df[local_logp_columns].max(axis=1, skipna=True)
    df["pass_local_filter"] = df["best_local_logp"] > args.local_ancestry_min_logp

    for ancestry in ANCESTRIES:
        df[f"signal_{ancestry}"] = (
            df["pass_local_filter"]
            & (df[f"p_dosage_{ancestry}"] <= args.dosage_gwas_max_p)
        )

    signal_columns = [f"signal_{ancestry}" for ancestry in ANCESTRIES]
    df["n_signal_ancestries"] = df[signal_columns].sum(axis=1)
    df["any_dosage_signal"] = df[signal_columns].any(axis=1)
    df["any_local_support"] = df["pass_local_filter"]
    df["passes_report_filters"] = (
        df["pass_cohort_maf"]
        & df["pass_ancestry_maf"]
        & df["pass_local_filter"]
    )

    df["best_p_dosage_any"] = df[[f"p_dosage_{ancestry}" for ancestry in ANCESTRIES]].min(axis=1, skipna=True)
    df["best_ancestry"] = df.apply(best_ancestry, axis=1)
    df["signal_pattern"] = df.apply(
        lambda row: classify_active_ancestries([ancestry for ancestry in ANCESTRIES if row[f"signal_{ancestry}"]]),
        axis=1,
    )
    return df


def pick_lead_row(group: pd.DataFrame, ancestry: str | None = None) -> pd.Series | None:
    if group.empty:
        return None
    if ancestry is None:
        candidates = group.dropna(subset=["best_p_dosage_any"]).sort_values(["best_p_dosage_any", "pos", "rsid"])
    else:
        column = f"p_dosage_{ancestry}"
        candidates = group.dropna(subset=[column]).sort_values([column, "pos", "rsid"])
    if candidates.empty:
        return None
    return candidates.iloc[0]


def build_locus_report(df: pd.DataFrame, min_supporting_snps: int) -> pd.DataFrame:
    records = []

    for (chrom, locus_name), group in df.groupby(["chr", "locus_name"], sort=False):
        group_all = group.sort_values(["pos", "rsid"]).copy()
        group_filtered = group_all[group_all["passes_report_filters"]].copy()
        lead_pool = group_filtered if not group_filtered.empty else group_all

        signal_counts = {
            ancestry: int(group_filtered[f"signal_{ancestry}"].sum())
            for ancestry in ANCESTRIES
        }
        active = [ancestry for ancestry in ANCESTRIES if signal_counts[ancestry] >= min_supporting_snps]
        classification = classify_active_ancestries(active)

        lead_any = pick_lead_row(lead_pool)
        genes_used = summarise_genes(lead_pool["gene"] if not lead_pool.empty else group_all["gene"])
        start_pos = int(group_all["pos"].min())
        end_pos = int(group_all["pos"].max())

        record = {
            "chr": chrom,
            "locus_name": locus_name,
            "region": format_region(chrom, start_pos, end_pos),
            "genes": genes_used,
            "start_pos": start_pos,
            "end_pos": end_pos,
            "n_snps_total": int(len(group_all)),
            "n_snps_passing_filters": int(len(group_filtered)),
            "n_local_support_snps": int(group_filtered["any_local_support"].sum()),
            "active_ancestries": ",".join(active) if active else "none",
            "n_active_ancestries": len(active),
            "classification": classification,
            "classification_label": CLASSIFICATION_LABELS.get(classification, classification),
            "classification_note": CLASSIFICATION_NOTES.get(classification, ""),
            "lead_rsid": lead_any["rsid"] if lead_any is not None else "",
            "lead_gene": lead_any["gene"] if lead_any is not None else "",
            "lead_pos": int(lead_any["pos"]) if lead_any is not None and pd.notna(lead_any["pos"]) else "",
            "lead_best_ancestry": lead_any["best_ancestry"] if lead_any is not None else "",
            "lead_best_p_dosage": lead_any["best_p_dosage_any"] if lead_any is not None else np.nan,
        }

        for ancestry in ANCESTRIES:
            lead_ancestry = pick_lead_row(group_filtered[group_filtered[f"signal_{ancestry}"]], ancestry)
            if lead_ancestry is None:
                lead_ancestry = pick_lead_row(lead_pool, ancestry)

            record[f"n_signal_snps_{ancestry}"] = signal_counts[ancestry]
            record[f"lead_rsid_{ancestry}"] = lead_ancestry["rsid"] if lead_ancestry is not None else ""
            record[f"best_p_dosage_{ancestry}"] = lead_ancestry[f"p_dosage_{ancestry}"] if lead_ancestry is not None else np.nan
            record[f"best_OR_dosage_{ancestry}"] = lead_ancestry[f"OR_dosage_{ancestry}"] if lead_ancestry is not None else np.nan

        record["best_p_local_ancestry"] = (
            group_filtered[[f"p_local_ancestry_{pair}" for pair in LOCAL_PAIRS]].min(axis=1, skipna=True).min(skipna=True)
            if not group_filtered.empty else np.nan
        )
        record["best_local_logp"] = group_filtered["best_local_logp"].max(skipna=True) if not group_filtered.empty else np.nan
        records.append(record)

    report = pd.DataFrame.from_records(records)
    if report.empty:
        return report

    chr_rank, chr_label = chromosome_sort_parts(report["chr"])
    report["__chr_rank"] = chr_rank
    report["__chr_label"] = chr_label
    report = report.sort_values(["__chr_rank", "__chr_label", "start_pos", "locus_name"]).drop(columns=["__chr_rank", "__chr_label"])
    return report


def build_parameters_table(args) -> pd.DataFrame:
    rows = [
        ("cohort_maf_min", args.cohort_maf_min, "Keep SNPs with cohort MAF greater than or equal to this value."),
        ("ancestry_maf_min", args.ancestry_maf_min, "Require at least one ancestry-specific MAF to meet this minimum."),
        ("local_ancestry_min_logp", args.local_ancestry_min_logp, "Filter SNPs first by requiring -log10(local ancestry p-value) to be above this threshold."),
        ("dosage_gwas_max_p", args.dosage_gwas_max_p, "Among SNPs that pass the local ancestry filter, require ancestry-specific GWAS dosage p-value at or below this threshold."),
        ("min_supporting_snps", args.min_supporting_snps, "Minimum number of supporting SNPs to mark an ancestry as active for a locus."),
        ("clustering_method", args.clustering_method, "Hierarchical clustering linkage method used for the dendrogram."),
        ("clustering_metric", args.clustering_metric, "Distance metric used to compare loci in the dendrogram."),
    ]
    return pd.DataFrame(rows, columns=["parameter", "value", "description"])


def build_conclusions_table(report: pd.DataFrame) -> pd.DataFrame:
    total = len(report)
    if total == 0:
        return pd.DataFrame([
            ("summary", "total_loci", 0),
            ("summary", "message", "No loci passed into the report."),
        ], columns=["section", "item", "value"])

    counts = report["classification"].value_counts().to_dict()
    single_count = int((report["n_active_ancestries"] == 1).sum())
    shared_subset = int(((report["n_active_ancestries"] >= 2) & (report["n_active_ancestries"] < len(ANCESTRIES))).sum())
    shared_all = int((report["n_active_ancestries"] == len(ANCESTRIES)).sum())
    no_clear = int((report["n_active_ancestries"] == 0).sum())
    strongest = report.sort_values(["lead_best_p_dosage", "chr", "start_pos"]).head(5)
    most_common = report["classification_label"].value_counts().idxmax()

    rows = [
        ("summary", "total_loci", total),
        ("summary", "one_ancestry_loci", single_count),
        ("summary", "shared_subset_loci", shared_subset),
        ("summary", "shared_all_ancestries_loci", shared_all),
        ("summary", "no_clear_signal_loci", no_clear),
        ("summary", "most_common_pattern", most_common),
    ]
    for key, value in counts.items():
        rows.append(("classification_counts", CLASSIFICATION_LABELS.get(key, key), int(value)))

    for _, row in strongest.iterrows():
        rows.append((
            "top_loci",
            f"{row['locus_name']} | {row['lead_gene']} | {row['region']}",
            CLASSIFICATION_LABELS.get(row["classification"], row["classification"]),
        ))

    return pd.DataFrame(rows, columns=["section", "item", "value"])


def build_conclusion_lines(report: pd.DataFrame) -> list[str]:
    if report.empty:
        return ["No loci were available for clustering."]

    total = len(report)
    single_count = int((report["n_active_ancestries"] == 1).sum())
    shared_subset = int(((report["n_active_ancestries"] >= 2) & (report["n_active_ancestries"] < len(ANCESTRIES))).sum())
    shared_all = int((report["n_active_ancestries"] == len(ANCESTRIES)).sum())
    no_clear = int((report["n_active_ancestries"] == 0).sum())
    most_common = report["classification_label"].value_counts().idxmax()
    strongest = report.sort_values(["lead_best_p_dosage", "chr", "start_pos"]).head(3)

    lines = [
        f"Total loci clustered: {total}",
        f"Single-ancestry loci: {single_count}",
        f"Shared across ancestry subsets: {shared_subset}",
        f"Shared across all ancestries: {shared_all}",
        f"No-clear-signal loci: {no_clear}",
        f"Most common pattern: {most_common}",
    ]
    if not strongest.empty:
        top_labels = [f"{row['locus_name']} ({row['lead_gene']})" for _, row in strongest.iterrows()]
        lines.append("Top lead loci: " + ", ".join(top_labels))
    return lines


def auto_fit_worksheet(ws):
    for column_cells in ws.columns:
        length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
            except Exception:
                value = ""
            length = max(length, len(value))
        ws.column_dimensions[column_letter].width = min(max(length + 2, 12), 42)


def style_sheet_headers(ws, freeze="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def apply_number_formats(ws):
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(row=1, column=cell.column).value
            if header is None:
                continue
            if header.startswith("best_p") or header.startswith("p_"):
                cell.number_format = "0.00E+00"
            elif header.startswith("best_OR") or header.startswith("OR_"):
                cell.number_format = "0.00"
            elif header.endswith("_maf"):
                cell.number_format = "0.000"
            elif header in {"start_pos", "end_pos", "lead_pos", "n_snps_total", "n_snps_passing_filters", "n_local_support_snps", "n_active_ancestries"} or str(header).startswith("n_signal_snps_"):
                cell.number_format = "#,##0"


def write_excel_report(locus_report: pd.DataFrame, filtered_snps: pd.DataFrame, parameters: pd.DataFrame, output_path: str):
    conclusions = build_conclusions_table(locus_report)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        locus_report.to_excel(writer, sheet_name="Classification", index=False)
        parameters.to_excel(writer, sheet_name="Parameters", index=False)
        conclusions.to_excel(writer, sheet_name="Conclusions", index=False)
        filtered_snps.to_excel(writer, sheet_name="Filtered_SNPs", index=False)

    wb = load_workbook(output_path)
    for name in ["Classification", "Parameters", "Conclusions", "Filtered_SNPs"]:
        ws = wb[name]
        style_sheet_headers(ws)
        auto_fit_worksheet(ws)
        apply_number_formats(ws)

    ws = wb["Classification"]
    class_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "classification":
            class_col = idx
            break
    if class_col is not None:
        for row_idx in range(2, ws.max_row + 1):
            classification = ws.cell(row=row_idx, column=class_col).value or "no_clear_signal"
            color = CLASSIFICATION_COLORS.get(classification, "#B8BEC8").replace("#", "")
            fill = PatternFill("solid", fgColor=color)
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
            ws.cell(row=row_idx, column=class_col).font = Font(bold=True, color="111111")

    wb["Parameters"].freeze_panes = "A2"
    wb["Conclusions"].freeze_panes = "A2"
    wb["Filtered_SNPs"].freeze_panes = "A2"
    wb.save(output_path)


def build_cluster_feature_matrix(report: pd.DataFrame) -> pd.DataFrame:
    feature_df = pd.DataFrame(index=report.index)
    feature_df["n_active_ancestries"] = pd.to_numeric(report["n_active_ancestries"], errors="coerce").fillna(0)
    feature_df["n_snps_passing_filters"] = pd.to_numeric(report["n_snps_passing_filters"], errors="coerce").fillna(0)
    feature_df["n_local_support_snps"] = pd.to_numeric(report["n_local_support_snps"], errors="coerce").fillna(0)

    total = pd.to_numeric(report["n_snps_total"], errors="coerce").replace(0, np.nan)
    feature_df["fraction_snps_passing"] = feature_df["n_snps_passing_filters"] / total

    for ancestry in ANCESTRIES:
        feature_df[f"n_signal_{ancestry}"] = pd.to_numeric(report[f"n_signal_snps_{ancestry}"], errors="coerce").fillna(0)
        pvals = pd.to_numeric(report[f"best_p_dosage_{ancestry}"], errors="coerce").clip(lower=1e-300)
        ors = pd.to_numeric(report[f"best_OR_dosage_{ancestry}"], errors="coerce").clip(lower=1e-12)
        feature_df[f"logp_{ancestry}"] = -np.log10(pvals.fillna(1.0))
        feature_df[f"abslogor_{ancestry}"] = np.abs(np.log(ors.fillna(1.0)))

    local_p = pd.to_numeric(report["best_p_local_ancestry"], errors="coerce").clip(lower=1e-300)
    feature_df["logp_local"] = -np.log10(local_p.fillna(1.0))
    feature_df = feature_df.replace([np.inf, -np.inf], np.nan)
    feature_df = feature_df.fillna(feature_df.median()).fillna(0.0)

    std = feature_df.std(ddof=0).replace(0, 1)
    return (feature_df - feature_df.mean()) / std


def render_dendrogram_svg(report: pd.DataFrame, output_path: str, method: str, metric: str):
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    if report.empty:
        output.write_text(
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="220"><text x="40" y="80" font-size="20">No loci available for clustering.</text></svg>'
        )
        return

    cluster_df = report.copy()
    cluster_df["cluster_label"] = cluster_df.apply(
        lambda row: f"{row['locus_name']} | {row['lead_gene'] or 'NA'} | {row['region']}", axis=1
    )

    if len(cluster_df) == 1:
        row = cluster_df.iloc[0]
        color = CLASSIFICATION_COLORS.get(row["classification"], "#B8BEC8")
        svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="1200" height="260">
  <rect width="1200" height="260" fill="white"/>
  <text x="30" y="40" font-size="24" font-weight="bold">Locus ancestry dendrogram</text>
  <text x="30" y="72" font-size="15">Only one locus passed into the clustering stage.</text>
  <rect x="30" y="105" width="24" height="24" fill="{color}" stroke="#333"/>
  <text x="66" y="122" font-size="16">{row['cluster_label']}</text>
  <text x="30" y="164" font-size="15">Conclusion: {CLASSIFICATION_NOTES.get(row['classification'], row['classification'])}</text>
</svg>'''
        output.write_text(svg)
        return

    features = build_cluster_feature_matrix(cluster_df)
    linkage_matrix = linkage(features.values, method=method, metric=metric)
    dendro = dendrogram(linkage_matrix, labels=cluster_df["cluster_label"].tolist(), no_plot=True)

    plot_left = 40
    plot_top = 130
    plot_width = 560
    right_label_space = 760
    row_spacing = 26
    plot_height = max(280, len(cluster_df) * row_spacing)
    width = plot_left + plot_width + right_label_space
    height = plot_top + plot_height + 60

    max_i = max(max(coords) for coords in dendro["icoord"]) if dendro["icoord"] else 10
    max_d = max(max(coords) for coords in dendro["dcoord"]) if dendro["dcoord"] else 1.0
    max_d = max(max_d, 1e-9)

    def x_scale(distance):
        return plot_left + ((max_d - distance) / max_d) * plot_width

    def y_scale(i_value):
        return plot_top + (i_value / max_i) * plot_height

    lines = []
    for icoord, dcoord in zip(dendro["icoord"], dendro["dcoord"]):
        points = [(x_scale(d), y_scale(i)) for i, d in zip(icoord, dcoord)]
        for (x1, y1), (x2, y2) in zip(points[:-1], points[1:]):
            lines.append(f'<line x1="{x1:.2f}" y1="{y1:.2f}" x2="{x2:.2f}" y2="{y2:.2f}" stroke="#8B95A7" stroke-width="2" />')

    leaf_y = {leaf_idx: y_scale(5 + 10 * order) for order, leaf_idx in enumerate(dendro["leaves"])}

    legend_items = []
    lx = plot_left
    ly = 28
    for idx, (key, label) in enumerate(CLASSIFICATION_LABELS.items()):
        x0 = lx + (idx % 3) * 280
        y0 = ly + (idx // 3) * 24
        legend_items.append(f'<rect x="{x0}" y="{y0}" width="14" height="14" fill="{CLASSIFICATION_COLORS[key]}" stroke="#333" />')
        legend_items.append(f'<text x="{x0 + 20}" y="{y0 + 12}" font-size="12">{label}</text>')

    leaf_labels = []
    for leaf_idx in dendro["leaves"]:
        row = cluster_df.iloc[leaf_idx]
        y = leaf_y[leaf_idx]
        color = CLASSIFICATION_COLORS.get(row["classification"], "#B8BEC8")
        leaf_labels.append(f'<circle cx="{plot_left + plot_width + 4}" cy="{y:.2f}" r="5" fill="{color}" stroke="#333" />')
        leaf_labels.append(f'<text x="{plot_left + plot_width + 16}" y="{y + 4:.2f}" font-size="12.5" fill="{color}" font-weight="bold">{row["cluster_label"]}</text>')
        leaf_labels.append(f'<text x="{plot_left + plot_width + 16}" y="{y + 18:.2f}" font-size="11" fill="#445066">{CLASSIFICATION_LABELS.get(row["classification"], row["classification"])} | active={row["active_ancestries"]}</text>')

    conclusion_lines = build_conclusion_lines(cluster_df)
    conclusion_svg = []
    box_x = plot_left
    box_y = plot_top + plot_height + 14
    box_h = 20 + len(conclusion_lines) * 16
    conclusion_svg.append(f'<rect x="{box_x}" y="{box_y}" width="{width - 2 * plot_left}" height="{box_h}" rx="10" ry="10" fill="#F8FAFD" stroke="#CBD5E1" />')
    conclusion_svg.append(f'<text x="{box_x + 14}" y="{box_y + 20}" font-size="14" font-weight="bold">Conclusions</text>')
    for idx, line in enumerate(conclusion_lines):
        conclusion_svg.append(f'<text x="{box_x + 18}" y="{box_y + 40 + idx * 16}" font-size="12">• {line}</text>')

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}">
  <rect width="{width}" height="{height}" fill="white"/>
  <text x="{plot_left}" y="20" font-size="24" font-weight="bold">Locus ancestry dendrogram</text>
  <text x="{plot_left}" y="42" font-size="13">Hierarchical clustering based on locus-level dosage/local ancestry signal profiles. Leaf labels show locus, lead gene, and region.</text>
  {''.join(legend_items)}
  <line x1="{plot_left}" y1="{plot_top + plot_height}" x2="{plot_left + plot_width}" y2="{plot_top + plot_height}" stroke="#CBD5E1" />
  {''.join(lines)}
  {''.join(leaf_labels)}
  {''.join(conclusion_svg)}
</svg>'''
    output.write_text(svg)


def main():
    global ANCESTRIES, LOCAL_PAIRS, CLASSIFICATION_LABELS, CLASSIFICATION_NOTES, CLASSIFICATION_COLORS

    args = parse_args()
    df, detected_ancestries, detected_pairs = load_table(args.input)
    ANCESTRIES = detected_ancestries
    LOCAL_PAIRS = detected_pairs
    CLASSIFICATION_LABELS, CLASSIFICATION_NOTES, CLASSIFICATION_COLORS = build_classification_metadata(ANCESTRIES)
    df = add_filter_columns(df, args)

    locus_report = build_locus_report(df, args.min_supporting_snps)
    filtered_snps = df[df["passes_report_filters"]].copy()
    chr_rank, chr_label = chromosome_sort_parts(filtered_snps["chr"] if not filtered_snps.empty else pd.Series(dtype=str))
    if not filtered_snps.empty:
        filtered_snps["__chr_rank"] = chr_rank
        filtered_snps["__chr_label"] = chr_label
        filtered_snps = filtered_snps.sort_values(["__chr_rank", "__chr_label", "pos", "rsid"]).drop(columns=["__chr_rank", "__chr_label"])

    snp_columns = [
        "chr", "pos", "rsid", "locus_name", "gene", "ref", "alt",
        "cohort_maf",
    ]
    snp_columns.extend([f"{ancestry}_maf" for ancestry in ANCESTRIES])
    snp_columns.extend([f"p_dosage_{ancestry}" for ancestry in ANCESTRIES])
    snp_columns.extend([f"OR_dosage_{ancestry}" for ancestry in ANCESTRIES])
    snp_columns.extend([f"p_local_ancestry_{pair}" for pair in LOCAL_PAIRS])
    snp_columns.extend([f"OR_local_ancestry_{pair}" for pair in LOCAL_PAIRS])
    snp_columns.extend([f"local_logp_{pair}" for pair in LOCAL_PAIRS])
    snp_columns.extend([f"signal_{ancestry}" for ancestry in ANCESTRIES])
    snp_columns.extend(["pass_local_filter", "best_local_logp", "n_signal_ancestries", "signal_pattern", "best_ancestry"])
    filtered_snps = filtered_snps[snp_columns] if not filtered_snps.empty else pd.DataFrame(columns=snp_columns)

    parameters = build_parameters_table(args)

    output_locus = Path(args.output_locus)
    output_snp = Path(args.output_snp)
    output_excel = Path(args.output_excel)
    output_dendrogram = Path(args.output_dendrogram)
    output_locus.parent.mkdir(parents=True, exist_ok=True)
    output_snp.parent.mkdir(parents=True, exist_ok=True)
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    output_dendrogram.parent.mkdir(parents=True, exist_ok=True)

    locus_report.to_csv(output_locus, sep="\t", index=False, na_rep="")
    filtered_snps.to_csv(output_snp, sep="\t", index=False, na_rep="")
    write_excel_report(locus_report, filtered_snps, parameters, str(output_excel))
    render_dendrogram_svg(locus_report, str(output_dendrogram), args.clustering_method, args.clustering_metric)

    print(
        f"Wrote locus ancestry report to {output_locus} with {len(locus_report)} loci; "
        f"filtered SNP table to {output_snp} with {len(filtered_snps)} SNPs; "
        f"Excel workbook to {output_excel}; dendrogram to {output_dendrogram}"
    )


if __name__ == "__main__":
    main()
