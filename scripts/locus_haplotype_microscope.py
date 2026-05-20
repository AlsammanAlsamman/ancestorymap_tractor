#!/usr/bin/env python3
import argparse
import gzip
import math
import os
import re
import sys
from collections import OrderedDict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

try:
    from cyvcf2 import VCF
except Exception:
    VCF = None


TRACT_COLORS = {
    "AFR": "#EF4444",
    "AMR": "#F59E0B",
    "EAS": "#10B981",
    "EUR": "#3B82F6",
}
DEFAULT_COLOR = "#9CA3AF"
OR_HIGH_FILL = PatternFill(fill_type="solid", fgColor="FDE68A")
OR_LOW_FILL = PatternFill(fill_type="solid", fgColor="BFDBFE")
P_BEST_FILL = PatternFill(fill_type="solid", fgColor="FED7AA")
BASE_FONT = Font(name="Calibri", size=11)
HIGHLIGHT_FONT = Font(name="Calibri", size=11, bold=True)
MISSING_VALUE = np.nan


def warn(message):
    print("[locus_haplotype_microscope] {}".format(message), file=sys.stderr)


def sanitize_name(value):
    text = re.sub(r"[^A-Za-z0-9_]+", "_", str(value).strip())
    text = text.strip("_")
    return text or "sheet"


def resolve_config_path(path):
    if os.path.exists(path):
        return path
    if os.path.basename(path) == "analysis.yml":
        alt = os.path.join("configs", "analysis.yml")
        if os.path.exists(alt):
            return alt
    raise FileNotFoundError("Missing config file: {}".format(path))


def load_yaml(path):
    with open(path, "r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle) or {}
    if not isinstance(data, dict):
        raise ValueError("Config root must be a mapping: {}".format(path))
    return data


def infer_results_dir(cfg):
    explicit = cfg.get("results_dir")
    if explicit:
        return str(explicit)

    project_name = str((cfg.get("metadata") or {}).get("project_name", "")).strip()
    if "analysis-" in project_name.lower():
        suffix = project_name.split("Analysis-", 1)[-1] if "Analysis-" in project_name else project_name.split("analysis-", 1)[-1]
        suffix = re.sub(r"[^A-Za-z0-9]+", "_", suffix).strip("_").lower()
        if suffix:
            return "results_{}".format(suffix)

    cohort_name = str((cfg.get("cohort") or {}).get("name", "cohort")).strip().lower()
    cohort_name = re.sub(r"[^a-z0-9]+", "_", cohort_name).strip("_") or "cohort"
    return "results_{}".format(cohort_name)


def derive_runtime_config(cfg, regions_arg):
    results_dir = infer_results_dir(cfg)
    cohort_cfg = cfg.get("cohort") or {}
    ancestry_cfg = cfg.get("ancestry") or {}
    reference_cfg = cfg.get("reference") or {}
    data_cfg = cfg.get("data") or {}
    tractor_gwas_cfg = cfg.get("tractor_gwas") or {}

    plink_prefix = str(cohort_cfg.get("plink_prefix", "Hispanic"))
    fam_path = plink_prefix if "/" in plink_prefix else os.path.join("inputs", plink_prefix + ".fam")
    if not fam_path.endswith(".fam"):
        fam_path = fam_path + ".fam"

    populations = [str(item) for item in reference_cfg.get("populations", [])]
    pop_to_label = ancestry_cfg.get("population_to_label") or {}
    derived_labels = []
    for pop in populations:
        derived_labels.append(str(pop_to_label.get(pop, pop)))

    ancestry_labels = tractor_gwas_cfg.get("ancestry_labels")
    if isinstance(ancestry_labels, dict) and ancestry_labels:
        code_to_label = {str(code): str(label) for code, label in ancestry_labels.items()}
    else:
        code_to_label = {str(idx): label for idx, label in enumerate(derived_labels)}

    label_order = [code_to_label[key] for key in sorted(code_to_label, key=lambda value: int(value))]
    case_code = str(tractor_gwas_cfg.get("fam_case_code", cohort_cfg.get("case_code", 2)))
    control_code = str(tractor_gwas_cfg.get("fam_control_code", cohort_cfg.get("control_code", 1)))

    config_regions = cfg.get("validation_regions") or regions_arg
    if not config_regions:
        raise ValueError("validation_regions is missing from config and not provided on the command line")

    return {
        "results_dir": results_dir,
        "fam_path": fam_path,
        "step4_output_dir": os.path.join(results_dir, "cohort_phasing"),
        "step5_output_dir": os.path.join(results_dir, "rfmix"),
        "tractor_output_dir": os.path.join(results_dir, "tractor"),
        "ancestry_gwas_dir": os.path.join(results_dir, "ancestry_plink_gwas"),
        "code_to_label": code_to_label,
        "ancestry_order": label_order,
        "case_code": case_code,
        "control_code": control_code,
        "regions_file": str(config_regions),
        "plink_prefix": plink_prefix,
        "cohort_name": str(cohort_cfg.get("name", "cohort")),
        "loci_file": data_cfg.get("loci_file", ""),
    }


def parse_args():
    parser = argparse.ArgumentParser(description="Build per-locus haplotype microscope outputs")
    parser.add_argument("--config", required=True)
    parser.add_argument("--regions", required=True)
    parser.add_argument("--locus")
    parser.add_argument("--outdir", required=True)
    parser.add_argument("--n-samples", type=int, default=100)
    parser.add_argument("--merge-only", action="store_true")
    parser.add_argument("--counts-dir")
    args = parser.parse_args()

    if args.merge_only:
        if not args.counts_dir:
            parser.error("--counts-dir is required with --merge-only")
    elif not args.locus:
        parser.error("--locus is required unless --merge-only is used")
    return args


def open_text(path):
    if str(path).endswith(".gz"):
        return gzip.open(path, "rt")
    return open(path, "r", encoding="utf-8")


def extract_base_iid(sample_id):
    text = str(sample_id)
    match = re.match(r"^(.*?)(?:[._-](?:0|1|hap[12]|h[12]))$", text)
    if match:
        return match.group(1)
    underscores = [idx for idx, char in enumerate(text) if char == "_"]
    for idx in underscores:
        left = text[:idx]
        right = text[idx + 1 :]
        if left == right:
            return left
    return text


def normalize_column_name(name):
    return re.sub(r"[^A-Z0-9]+", "", str(name).upper())


def safe_float(value):
    try:
        result = float(value)
    except Exception:
        return MISSING_VALUE
    if math.isnan(result):
        return MISSING_VALUE
    return result


def safe_int(value):
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        return int(float(value))
    except Exception:
        return None


def safe_or(case_alt, case_ref, control_alt, control_ref):
    values = [case_alt, case_ref, control_alt, control_ref]
    if any(pd.isna(value) for value in values):
        return MISSING_VALUE
    if case_alt <= 0 or case_ref <= 0 or control_alt <= 0 or control_ref <= 0:
        return MISSING_VALUE
    return (float(case_alt) / float(case_ref)) / (float(control_alt) / float(control_ref))


def safe_freq(alt_count, total_count):
    if pd.isna(alt_count) or pd.isna(total_count) or float(total_count) <= 0:
        return MISSING_VALUE
    return float(alt_count) / float(total_count)


def load_fam_status(fam_path, case_code, control_code):
    fam = pd.read_csv(fam_path, sep=r"\s+", header=None, dtype=str, engine="python")
    if fam.shape[1] < 6:
        raise ValueError("FAM file must have at least 6 columns: {}".format(fam_path))
    fam.columns = ["FID", "IID", "PAT", "MAT", "SEX", "PHENO"]
    fam["IID"] = fam["IID"].astype(str)
    fam["BASE_IID"] = fam["IID"].map(extract_base_iid)

    case_ids = set(fam.loc[fam["PHENO"].astype(str) == str(case_code), "IID"].astype(str))
    case_ids.update(fam.loc[fam["PHENO"].astype(str) == str(case_code), "BASE_IID"].astype(str))
    control_ids = set(fam.loc[fam["PHENO"].astype(str) == str(control_code), "IID"].astype(str))
    control_ids.update(fam.loc[fam["PHENO"].astype(str) == str(control_code), "BASE_IID"].astype(str))
    return fam, case_ids, control_ids


def load_regions(regions_path):
    regions = pd.read_csv(regions_path, sep="\t", dtype=str)
    regions.columns = [str(col).strip() for col in regions.columns]
    norm = {normalize_column_name(col): col for col in regions.columns}
    if "LOCUS" not in norm:
        raise ValueError("Regions file must contain a 'locus' column: {}".format(regions_path))
    return regions, norm


def get_region_row(regions, norm_cols, locus):
    locus_col = norm_cols["LOCUS"]
    subset = regions.loc[regions[locus_col].astype(str) == str(locus)].copy()
    if subset.empty:
        raise ValueError("Locus '{}' was not found in {}".format(locus, locus_col))
    row = subset.iloc[0]
    chr_col = norm_cols.get("CHR") or norm_cols.get("CHROM") or norm_cols.get("CHROMOSOME")
    start_col = norm_cols.get("START") or norm_cols.get("STARTPOS")
    end_col = norm_cols.get("END") or norm_cols.get("ENDPOS") or norm_cols.get("STOP")
    gene_col = norm_cols.get("GENE")
    ancestry_class_col = norm_cols.get("ANCESTRYCLASS")
    if not chr_col or not start_col or not end_col:
        raise ValueError("Regions file requires chr/start/end columns")
    return {
        "locus": str(row[locus_col]),
        "chr": str(row[chr_col]).replace("chr", ""),
        "start": safe_int(row[start_col]),
        "end": safe_int(row[end_col]),
        "gene": str(row[gene_col]) if gene_col and pd.notna(row[gene_col]) else "",
        "ancestry_class": str(row[ancestry_class_col]) if ancestry_class_col and pd.notna(row[ancestry_class_col]) else "",
    }


def resolve_gwas_path(base_dir, label):
    return os.path.join(base_dir, label, "{}.gwas.with_case_control_freq.tsv".format(label))


def read_gwas_table(path):
    table = pd.read_csv(path, sep="\t", dtype=str)
    normalized = {normalize_column_name(col): col for col in table.columns}
    pos_col = None
    for candidate in ["POS", "BP", "POSITION"]:
        if candidate in normalized:
            pos_col = normalized[candidate]
            break
    p_col = None
    for candidate in ["P", "PVALUE", "PVALUE", "PVAL", "PVALUEADJ", "PVALUEGWAS", "P_VALUE"]:
        if candidate in normalized:
            p_col = normalized[candidate]
            break
    id_col = None
    for candidate in ["ID", "SNP", "VARIANTID"]:
        if candidate in normalized:
            id_col = normalized[candidate]
            break
    if pos_col is None or p_col is None or id_col is None:
        raise ValueError("GWAS file {} is missing required position/p-value/id columns".format(path))
    table = table.copy()
    table["_POS"] = pd.to_numeric(table[pos_col], errors="coerce")
    table["_P"] = pd.to_numeric(table[p_col], errors="coerce")
    table["_ID"] = table[id_col].astype(str)
    table["_REF"] = table[normalized["REF"]].astype(str) if "REF" in normalized else np.nan
    if "ALT" in normalized:
        table["_ALT"] = table[normalized["ALT"]].astype(str)
    elif "A1" in normalized:
        table["_ALT"] = table[normalized["A1"]].astype(str)
    else:
        table["_ALT"] = np.nan
    if "A2" in normalized and "REF" not in normalized:
        table["_REF"] = table[normalized["A2"]].astype(str)
    return table


def find_lead_snp(region, ancestry_order, gwas_dir):
    best = None
    per_ancestry = OrderedDict()

    for label in ancestry_order:
        path = resolve_gwas_path(gwas_dir, label)
        if not os.path.exists(path):
            warn("Missing GWAS file for ancestry {}: {}".format(label, path))
            per_ancestry[label] = None
            continue
        try:
            table = read_gwas_table(path)
        except Exception as exc:
            warn("Failed to read GWAS file {}: {}".format(path, exc))
            per_ancestry[label] = None
            continue

        window = table.loc[(table["_POS"] >= region["start"]) & (table["_POS"] <= region["end"])].copy()
        window = window.loc[window["_P"].notna()].sort_values(["_P", "_POS", "_ID"])
        if window.empty:
            per_ancestry[label] = None
            continue

        lead_row = window.iloc[0]
        per_ancestry[label] = lead_row
        candidate = {
            "ancestry": label,
            "p": safe_float(lead_row["_P"]),
            "id": str(lead_row["_ID"]),
            "pos": safe_int(lead_row["_POS"]),
            "ref": str(lead_row["_REF"]) if pd.notna(lead_row["_REF"]) else "",
            "alt": str(lead_row["_ALT"]) if pd.notna(lead_row["_ALT"]) else "",
        }
        if best is None or candidate["p"] < best["p"]:
            best = candidate

    return best, per_ancestry


def find_gwas_value_for_lead(path, lead_id, lead_pos):
    if not os.path.exists(path):
        return None
    try:
        table = read_gwas_table(path)
    except Exception as exc:
        warn("Failed to read GWAS file {}: {}".format(path, exc))
        return None

    by_id = table.loc[table["_ID"].astype(str) == str(lead_id)]
    if not by_id.empty:
        return by_id.iloc[0]
    by_pos = table.loc[table["_POS"] == int(lead_pos)]
    if not by_pos.empty:
        return by_pos.iloc[0]
    return None


def detect_sample_start(columns):
    fixed = {"#CHROM", "CHROM", "POS", "ID", "REF", "ALT", "QUAL", "FILTER", "INFO", "FORMAT"}
    for idx, column in enumerate(columns):
        if idx >= 5 and normalize_column_name(column) not in fixed:
            return idx
    return 5 if len(columns) > 5 else len(columns)


def find_variant_row(path, lead_id, lead_pos):
    if not os.path.exists(path):
        warn("Missing variant table: {}".format(path))
        return None

    with open_text(path) as handle:
        header = handle.readline().rstrip("\n").split("\t")
        sample_start = detect_sample_start(header)
        sample_ids = header[sample_start:]
        pos_idx = 1 if len(header) > 1 else None
        id_idx = 2 if len(header) > 2 else None
        ref_idx = 3 if len(header) > 3 else None
        alt_idx = 4 if len(header) > 4 else None

        by_pos = None
        for raw_line in handle:
            parts = raw_line.rstrip("\n").split("\t")
            if len(parts) < sample_start:
                continue
            snp_id = str(parts[id_idx]) if id_idx is not None and id_idx < len(parts) else ""
            pos_value = safe_int(parts[pos_idx]) if pos_idx is not None and pos_idx < len(parts) else None
            if snp_id == str(lead_id):
                return {
                    "chrom": str(parts[0]),
                    "pos": pos_value,
                    "id": snp_id,
                    "ref": str(parts[ref_idx]) if ref_idx is not None and ref_idx < len(parts) else "",
                    "alt": str(parts[alt_idx]) if alt_idx is not None and alt_idx < len(parts) else "",
                    "sample_ids": sample_ids,
                    "values": np.asarray(parts[sample_start:], dtype=float),
                }
            if by_pos is None and pos_value == int(lead_pos):
                by_pos = {
                    "chrom": str(parts[0]),
                    "pos": pos_value,
                    "id": snp_id,
                    "ref": str(parts[ref_idx]) if ref_idx is not None and ref_idx < len(parts) else "",
                    "alt": str(parts[alt_idx]) if alt_idx is not None and alt_idx < len(parts) else "",
                    "sample_ids": sample_ids,
                    "values": np.asarray(parts[sample_start:], dtype=float),
                }
        return by_pos


def build_group_mask(sample_ids, case_ids, control_ids):
    labels = []
    for sample_id in sample_ids:
        base_id = extract_base_iid(sample_id)
        if sample_id in case_ids or base_id in case_ids:
            labels.append("case")
        elif sample_id in control_ids or base_id in control_ids:
            labels.append("control")
        else:
            labels.append("other")
    labels = np.asarray(labels)
    return labels == "case", labels == "control"


def dosage_distribution(values):
    rounded = np.asarray(np.rint(values), dtype=int)
    return {
        0: int((rounded == 0).sum()),
        1: int((rounded == 1).sum()),
        2: int((rounded == 2).sum()),
    }


def compute_ancestry_metrics(label, dosage_row, hapcount_row, case_ids, control_ids):
    template = {
        "n_cases_with_haplotype": MISSING_VALUE,
        "n_controls_with_haplotype": MISSING_VALUE,
        "total_hap_cases": MISSING_VALUE,
        "total_hap_controls": MISSING_VALUE,
        "cases_alt_count": MISSING_VALUE,
        "cases_ref_count": MISSING_VALUE,
        "controls_alt_count": MISSING_VALUE,
        "controls_ref_count": MISSING_VALUE,
        "cases_alt_freq": MISSING_VALUE,
        "controls_alt_freq": MISSING_VALUE,
        "OR": MISSING_VALUE,
        "dosage0_cases": MISSING_VALUE,
        "dosage1_cases": MISSING_VALUE,
        "dosage2_cases": MISSING_VALUE,
        "dosage0_controls": MISSING_VALUE,
        "dosage1_controls": MISSING_VALUE,
        "dosage2_controls": MISSING_VALUE,
        "warning": "",
    }
    warnings = []

    if dosage_row is None:
        warnings.append("missing_dosage")
    if hapcount_row is None:
        warnings.append("missing_hapcount")
    if dosage_row is None or hapcount_row is None:
        template["warning"] = ";".join(warnings)
        return template

    if list(dosage_row["sample_ids"]) != list(hapcount_row["sample_ids"]):
        warnings.append("sample_order_mismatch")
        template["warning"] = ";".join(warnings)
        return template

    case_mask, control_mask = build_group_mask(dosage_row["sample_ids"], case_ids, control_ids)
    dosage = np.asarray(dosage_row["values"], dtype=float)
    hapcount = np.asarray(hapcount_row["values"], dtype=float)

    case_dosage = dosage[case_mask]
    control_dosage = dosage[control_mask]
    case_hap = hapcount[case_mask]
    control_hap = hapcount[control_mask]

    total_hap_cases = float(np.nansum(case_hap)) if case_hap.size else MISSING_VALUE
    total_hap_controls = float(np.nansum(control_hap)) if control_hap.size else MISSING_VALUE
    cases_alt_count = float(np.nansum(case_dosage)) if case_dosage.size else MISSING_VALUE
    controls_alt_count = float(np.nansum(control_dosage)) if control_dosage.size else MISSING_VALUE
    cases_ref_count = total_hap_cases - cases_alt_count if not pd.isna(total_hap_cases) and not pd.isna(cases_alt_count) else MISSING_VALUE
    controls_ref_count = total_hap_controls - controls_alt_count if not pd.isna(total_hap_controls) and not pd.isna(controls_alt_count) else MISSING_VALUE

    if not pd.isna(total_hap_cases) and total_hap_cases == 0:
        warnings.append("zero_haplotypes_cases")
    if not pd.isna(total_hap_controls) and total_hap_controls == 0:
        warnings.append("zero_haplotypes_controls")

    case_dist = dosage_distribution(case_dosage) if case_dosage.size else {0: MISSING_VALUE, 1: MISSING_VALUE, 2: MISSING_VALUE}
    control_dist = dosage_distribution(control_dosage) if control_dosage.size else {0: MISSING_VALUE, 1: MISSING_VALUE, 2: MISSING_VALUE}

    template.update(
        {
            "n_cases_with_haplotype": int((case_hap > 0).sum()) if case_hap.size else MISSING_VALUE,
            "n_controls_with_haplotype": int((control_hap > 0).sum()) if control_hap.size else MISSING_VALUE,
            "total_hap_cases": total_hap_cases,
            "total_hap_controls": total_hap_controls,
            "cases_alt_count": cases_alt_count,
            "cases_ref_count": cases_ref_count,
            "controls_alt_count": controls_alt_count,
            "controls_ref_count": controls_ref_count,
            "cases_alt_freq": safe_freq(cases_alt_count, total_hap_cases),
            "controls_alt_freq": safe_freq(controls_alt_count, total_hap_controls),
            "OR": safe_or(cases_alt_count, cases_ref_count, controls_alt_count, controls_ref_count),
            "dosage0_cases": case_dist[0],
            "dosage1_cases": case_dist[1],
            "dosage2_cases": case_dist[2],
            "dosage0_controls": control_dist[0],
            "dosage1_controls": control_dist[1],
            "dosage2_controls": control_dist[2],
            "warning": ";".join(warnings),
        }
    )
    return template


def compute_whole_cohort_from_vcf(vcf_path, chrom, lead_id, lead_pos, case_ids, control_ids):
    metrics = {
        "cases_ref_hom": MISSING_VALUE,
        "cases_het": MISSING_VALUE,
        "cases_alt_hom": MISSING_VALUE,
        "cases_alt_count": MISSING_VALUE,
        "cases_alt_freq": MISSING_VALUE,
        "controls_ref_hom": MISSING_VALUE,
        "controls_het": MISSING_VALUE,
        "controls_alt_hom": MISSING_VALUE,
        "controls_alt_count": MISSING_VALUE,
        "controls_alt_freq": MISSING_VALUE,
        "OR_whole_cohort": MISSING_VALUE,
        "n_cases_total": MISSING_VALUE,
        "n_controls_total": MISSING_VALUE,
        "ref": "",
        "alt": "",
    }

    if VCF is None:
        warn("cyvcf2 is unavailable; falling back to reconstructed whole-cohort dosage")
        return None
    if not os.path.exists(vcf_path):
        warn("Missing phased VCF: {}".format(vcf_path))
        return None

    try:
        reader = VCF(vcf_path)
    except Exception as exc:
        warn("Failed to open phased VCF {}: {}".format(vcf_path, exc))
        return None

    samples = list(reader.samples)
    case_mask = []
    control_mask = []
    for sample_id in samples:
        base_id = extract_base_iid(sample_id)
        case_mask.append(sample_id in case_ids or base_id in case_ids)
        control_mask.append(sample_id in control_ids or base_id in control_ids)
    case_mask = np.asarray(case_mask)
    control_mask = np.asarray(control_mask)

    record = None
    query_strings = [
        "{}:{}-{}".format(chrom, lead_pos, lead_pos),
        "chr{}:{}-{}".format(chrom, lead_pos, lead_pos),
    ]
    for query in query_strings:
        try:
            for item in reader(query):
                if str(item.ID) == str(lead_id) or int(item.POS) == int(lead_pos):
                    record = item
                    break
        except Exception:
            continue
        if record is not None:
            break

    if record is None:
        try:
            for item in reader:
                if str(item.ID) == str(lead_id) or int(item.POS) == int(lead_pos):
                    record = item
                    break
        except Exception as exc:
            warn("Failed while scanning phased VCF {}: {}".format(vcf_path, exc))
            return None

    if record is None:
        warn("Lead SNP was not found in phased VCF: {} ({})".format(vcf_path, lead_id))
        return None

    genotypes = np.asarray(record.genotypes)[:, :2]
    dosage = genotypes.sum(axis=1)
    dosage = np.where((genotypes < 0).any(axis=1), np.nan, dosage)

    case_dosage = dosage[case_mask]
    control_dosage = dosage[control_mask]
    case_valid = case_dosage[~np.isnan(case_dosage)]
    control_valid = control_dosage[~np.isnan(control_dosage)]

    case_dist = dosage_distribution(case_valid) if case_valid.size else {0: MISSING_VALUE, 1: MISSING_VALUE, 2: MISSING_VALUE}
    control_dist = dosage_distribution(control_valid) if control_valid.size else {0: MISSING_VALUE, 1: MISSING_VALUE, 2: MISSING_VALUE}
    case_alt = float(np.nansum(case_valid)) if case_valid.size else MISSING_VALUE
    control_alt = float(np.nansum(control_valid)) if control_valid.size else MISSING_VALUE
    case_total = float(2 * len(case_valid)) if case_valid.size else MISSING_VALUE
    control_total = float(2 * len(control_valid)) if control_valid.size else MISSING_VALUE
    case_ref = case_total - case_alt if not pd.isna(case_total) and not pd.isna(case_alt) else MISSING_VALUE
    control_ref = control_total - control_alt if not pd.isna(control_total) and not pd.isna(control_alt) else MISSING_VALUE

    metrics.update(
        {
            "cases_ref_hom": case_dist[0],
            "cases_het": case_dist[1],
            "cases_alt_hom": case_dist[2],
            "cases_alt_count": case_alt,
            "cases_alt_freq": safe_freq(case_alt, case_total),
            "controls_ref_hom": control_dist[0],
            "controls_het": control_dist[1],
            "controls_alt_hom": control_dist[2],
            "controls_alt_count": control_alt,
            "controls_alt_freq": safe_freq(control_alt, control_total),
            "OR_whole_cohort": safe_or(case_alt, case_ref, control_alt, control_ref),
            "n_cases_total": int(len(case_valid)),
            "n_controls_total": int(len(control_valid)),
            "ref": str(record.REF),
            "alt": str(record.ALT[0]) if getattr(record, "ALT", None) else "",
        }
    )
    return metrics


def compute_whole_cohort_from_reconstructed(ancestry_rows, case_ids, control_ids):
    sample_ids = None
    total_dosage = None
    for label in ancestry_rows:
        row = ancestry_rows[label].get("dosage_row")
        if row is None:
            continue
        if sample_ids is None:
            sample_ids = list(row["sample_ids"])
            total_dosage = np.zeros(len(sample_ids), dtype=float)
        if list(row["sample_ids"]) != sample_ids:
            warn("Sample order mismatch across ancestry dosage files; whole-cohort fallback may be incomplete")
            continue
        total_dosage += np.asarray(row["values"], dtype=float)

    if sample_ids is None:
        return None

    case_mask, control_mask = build_group_mask(sample_ids, case_ids, control_ids)
    case_vals = total_dosage[case_mask]
    control_vals = total_dosage[control_mask]
    case_dist = dosage_distribution(case_vals) if case_vals.size else {0: MISSING_VALUE, 1: MISSING_VALUE, 2: MISSING_VALUE}
    control_dist = dosage_distribution(control_vals) if control_vals.size else {0: MISSING_VALUE, 1: MISSING_VALUE, 2: MISSING_VALUE}
    case_alt = float(np.nansum(case_vals)) if case_vals.size else MISSING_VALUE
    control_alt = float(np.nansum(control_vals)) if control_vals.size else MISSING_VALUE
    case_total = float(2 * len(case_vals)) if case_vals.size else MISSING_VALUE
    control_total = float(2 * len(control_vals)) if control_vals.size else MISSING_VALUE
    case_ref = case_total - case_alt if not pd.isna(case_total) and not pd.isna(case_alt) else MISSING_VALUE
    control_ref = control_total - control_alt if not pd.isna(control_total) and not pd.isna(control_alt) else MISSING_VALUE
    return {
        "cases_ref_hom": case_dist[0],
        "cases_het": case_dist[1],
        "cases_alt_hom": case_dist[2],
        "cases_alt_count": case_alt,
        "cases_alt_freq": safe_freq(case_alt, case_total),
        "controls_ref_hom": control_dist[0],
        "controls_het": control_dist[1],
        "controls_alt_hom": control_dist[2],
        "controls_alt_count": control_alt,
        "controls_alt_freq": safe_freq(control_alt, control_total),
        "OR_whole_cohort": safe_or(case_alt, case_ref, control_alt, control_ref),
        "n_cases_total": int(case_vals.size),
        "n_controls_total": int(control_vals.size),
        "ref": "",
        "alt": "",
    }


def pick_target_ancestry(region, ancestry_order):
    ancestry_class = str(region.get("ancestry_class") or "").strip()
    if ancestry_class in ancestry_order:
        return ancestry_class
    if "AMR" in ancestry_order:
        return "AMR"
    return ancestry_order[0] if ancestry_order else ancestry_class


def parse_msp_sample_pairs(header_tokens, data_width):
    sample_tokens = header_tokens[7:] if len(header_tokens) - 7 == data_width - 6 else header_tokens[6:]
    if len(sample_tokens) != data_width - 6:
        sample_tokens = header_tokens[-(data_width - 6):]

    grouped = OrderedDict()
    for token in sample_tokens:
        base = re.sub(r"(?:[._-](?:0|1|hap[12]|h[12]))$", "", str(token))
        grouped.setdefault(base, []).append(token)

    sample_pairs = []
    for base, members in grouped.items():
        if len(members) >= 2:
            sample_pairs.append((base, members[0], members[1]))
    return sample_tokens, sample_pairs


def load_msp_windows(msp_path, chrom, start, end):
    if not os.path.exists(msp_path):
        warn("Missing MSP file: {}".format(msp_path))
        return [], []

    windows = []
    sample_pairs = []
    sample_token_index = None
    with open_text(msp_path) as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith("#chm"):
                header_tokens = line.lstrip("#").split()
                continue
            if line.startswith("#"):
                continue
            parts = line.split()
            if sample_token_index is None:
                sample_tokens, sample_pairs = parse_msp_sample_pairs(header_tokens, len(parts))
                sample_token_index = {token: idx for idx, token in enumerate(sample_tokens)}
            row_chr = str(parts[0]).replace("chr", "")
            row_start = safe_int(parts[1])
            row_end = safe_int(parts[2])
            if row_chr != str(chrom):
                continue
            if row_end is None or row_start is None:
                continue
            if row_end < start or row_start > end:
                continue
            calls = parts[6:]
            windows.append((row_start, row_end, calls, sample_token_index))
    return windows, sample_pairs


def save_message_plot(path, title, message):
    fig, ax = plt.subplots(figsize=(8, 3))
    ax.axis("off")
    ax.set_title(title)
    ax.text(0.5, 0.5, message, ha="center", va="center", wrap=True)
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def decode_msp_call(call_value, code_to_label):
    key = str(call_value)
    return str(code_to_label.get(key, key))


def make_tract_plot(path, region, lead_id, lead_pos, ancestry_order, code_to_label, msp_path, case_ids, control_ids, n_samples):
    title = "{} {}".format(region["locus"], region.get("gene") or "").strip()
    windows, sample_pairs = load_msp_windows(msp_path, region["chr"], region["start"], region["end"])
    if not windows or not sample_pairs:
        warn("No MSP windows overlap locus {}; skipping tract plot".format(region["locus"]))
        save_message_plot(path, title, "No overlapping MSP windows were available for this locus.")
        return

    target = pick_target_ancestry(region, ancestry_order)
    ranked_cases = []
    ranked_controls = []
    for base, hap_a, hap_b in sample_pairs:
        canonical_base = extract_base_iid(base)
        canonical_hap_a = extract_base_iid(hap_a)
        canonical_hap_b = extract_base_iid(hap_b)
        calls_a = []
        calls_b = []
        valid = True
        for _, _, calls, token_index in windows:
            if hap_a not in token_index or hap_b not in token_index:
                valid = False
                break
            calls_a.append(calls[token_index[hap_a]])
            calls_b.append(calls[token_index[hap_b]])
        if not valid or not calls_a:
            continue
        prop = (
            sum(decode_msp_call(call, code_to_label) == target for call in calls_a)
            + sum(decode_msp_call(call, code_to_label) == target for call in calls_b)
        ) / float(2 * len(calls_a))

        if (
            canonical_base in case_ids
            or canonical_hap_a in case_ids
            or canonical_hap_b in case_ids
            or base in case_ids
            or hap_a in case_ids
            or hap_b in case_ids
        ):
            ranked_cases.append((prop, base, hap_a, hap_b))
        elif (
            canonical_base in control_ids
            or canonical_hap_a in control_ids
            or canonical_hap_b in control_ids
            or base in control_ids
            or hap_a in control_ids
            or hap_b in control_ids
        ):
            ranked_controls.append((prop, base, hap_a, hap_b))

    ranked_cases.sort(key=lambda item: (-item[0], item[1]))
    ranked_controls.sort(key=lambda item: (-item[0], item[1]))
    chosen_cases = ranked_cases[: max(1, n_samples // 2)]
    chosen_controls = ranked_controls[: max(1, n_samples // 2)]
    selected = chosen_cases + chosen_controls
    if not selected:
        save_message_plot(path, title, "No case/control samples could be aligned to the MSP file.")
        return

    fig_height = max(6, 0.16 * len(selected) * 2)
    fig, ax = plt.subplots(figsize=(12, fig_height))
    for idx, (_, base, hap_a, hap_b) in enumerate(selected):
        y_top = (len(selected) - idx) * 2
        for row_start, row_end, calls, token_index in windows:
            for offset, token in enumerate([hap_a, hap_b]):
                label = decode_msp_call(calls[token_index[token]], code_to_label)
                color = TRACT_COLORS.get(label, DEFAULT_COLOR)
                ax.hlines(y_top - offset, row_start, row_end, colors=color, linewidth=2.5)
        ax.text(region["start"] - max(1, (region["end"] - region["start"]) * 0.01), y_top - 0.4, base, ha="right", va="center", fontsize=7)

    separator_y = (len(chosen_controls) * 2) + 1 if chosen_controls else None
    if separator_y is not None and chosen_cases and chosen_controls:
        ax.axhline(separator_y, color="#374151", linestyle="--", linewidth=1)
    ax.axvline(lead_pos, color="#111827", linestyle=":", linewidth=1.2)
    ax.text(lead_pos, ax.get_ylim()[1] if ax.get_ylim()[1] else 1, lead_id, rotation=90, va="bottom", ha="center", fontsize=8)
    legend_labels = []
    for label in ancestry_order:
        legend_labels.append(plt.Line2D([0], [0], color=TRACT_COLORS.get(label, DEFAULT_COLOR), lw=3, label=label))
    if legend_labels:
        ax.legend(handles=legend_labels, loc="upper right", frameon=False)
    ax.set_title(title)
    ax.set_xlabel("Genomic position")
    ax.set_ylabel("Samples (cases above controls)")
    ax.set_xlim(region["start"], region["end"])
    ax.set_yticks([])
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def make_ancestry_composition_plot(path, region, ancestry_order, metrics_by_ancestry):
    title = "{} ancestry composition".format(region["locus"])
    case_totals = []
    control_totals = []
    for label in ancestry_order:
        case_totals.append(safe_float(metrics_by_ancestry[label]["total_hap_cases"]))
        control_totals.append(safe_float(metrics_by_ancestry[label]["total_hap_controls"]))
    case_sum = np.nansum(case_totals)
    control_sum = np.nansum(control_totals)
    if case_sum <= 0 and control_sum <= 0:
        save_message_plot(path, title, "No ancestry-specific haplotype counts were available at the lead SNP.")
        return

    fig, ax = plt.subplots(figsize=(7, 5))
    bottoms = [0.0, 0.0]
    for idx, label in enumerate(ancestry_order):
        values = [case_totals[idx] / case_sum if case_sum > 0 else 0.0, control_totals[idx] / control_sum if control_sum > 0 else 0.0]
        bars = ax.bar([0, 1], values, bottom=bottoms, color=TRACT_COLORS.get(label, DEFAULT_COLOR), width=0.6, label=label)
        for bar, value, bottom in zip(bars, values, bottoms):
            if value >= 0.05:
                ax.text(bar.get_x() + bar.get_width() / 2.0, bottom + value / 2.0, "{:.0%}".format(value), ha="center", va="center", fontsize=9, color="white")
        bottoms = [bottoms[i] + values[i] for i in range(2)]
    ax.set_xticks([0, 1])
    ax.set_xticklabels(["Cases", "Controls"])
    ax.set_ylim(0, 1)
    ax.set_ylabel("Fraction of haplotypes")
    ax.set_title(title)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def make_allele_frequency_plot(path, region, ancestry_order, metrics_by_ancestry, whole_metrics):
    title = "{} alt allele frequency".format(region["locus"])
    groups = ancestry_order + ["WHOLE"]
    case_values = []
    control_values = []
    colors = []
    for label in ancestry_order:
        case_values.append(metrics_by_ancestry[label]["cases_alt_freq"])
        control_values.append(metrics_by_ancestry[label]["controls_alt_freq"])
        colors.append(TRACT_COLORS.get(label, DEFAULT_COLOR))
    case_values.append(whole_metrics.get("cases_alt_freq", MISSING_VALUE))
    control_values.append(whole_metrics.get("controls_alt_freq", MISSING_VALUE))
    colors.append("#4B5563")

    overall_freq = safe_freq(
        safe_float(whole_metrics.get("cases_alt_count", MISSING_VALUE)) + safe_float(whole_metrics.get("controls_alt_count", MISSING_VALUE)),
        2 * (safe_float(whole_metrics.get("n_cases_total", MISSING_VALUE)) + safe_float(whole_metrics.get("n_controls_total", MISSING_VALUE))),
    )

    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(groups))
    width = 0.36
    case_bars = ax.bar(x - width / 2.0, [0 if pd.isna(v) else v for v in case_values], width=width, color=colors)
    control_bars = ax.bar(x + width / 2.0, [0 if pd.isna(v) else v for v in control_values], width=width, color=colors, hatch="//", alpha=0.75)
    for bars, values in [(case_bars, case_values), (control_bars, control_values)]:
        for bar, value in zip(bars, values):
            if pd.isna(value):
                continue
            ax.text(bar.get_x() + bar.get_width() / 2.0, value + 0.02, "{:.2f}".format(value), ha="center", va="bottom", fontsize=8)
    if not pd.isna(overall_freq):
        ax.axhline(overall_freq, color="#111827", linestyle="--", linewidth=1)
    ax.set_xticks(x)
    ax.set_xticklabels(ancestry_order + ["Whole cohort"])
    ax.set_ylim(0, 1)
    ax.set_ylabel("Alt allele frequency")
    ax.set_title(title)
    legend_handles = [
        plt.Rectangle((0, 0), 1, 1, facecolor="#6B7280", label="Cases"),
        plt.Rectangle((0, 0), 1, 1, facecolor="#6B7280", hatch="//", alpha=0.75, label="Controls"),
    ]
    ax.legend(handles=legend_handles, frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def build_output_row(region, lead, whole_metrics, ancestry_order, metrics_by_ancestry, per_ancestry_p, note):
    row = OrderedDict()
    row["locus"] = region["locus"]
    row["gene"] = region.get("gene", "")
    row["chr"] = region["chr"]
    row["start"] = region["start"]
    row["end"] = region["end"]
    row["rsid"] = lead.get("id", "") if lead else ""
    row["pos"] = lead.get("pos", MISSING_VALUE) if lead else MISSING_VALUE
    row["ref"] = lead.get("ref", "") if lead else ""
    row["alt"] = lead.get("alt", "") if lead else ""
    row["lead_ancestry"] = lead.get("ancestry", "") if lead else ""
    row["lead_p"] = lead.get("p", MISSING_VALUE) if lead else MISSING_VALUE
    row["n_cases_total"] = whole_metrics.get("n_cases_total", MISSING_VALUE)
    row["n_controls_total"] = whole_metrics.get("n_controls_total", MISSING_VALUE)
    row["cases_ref_hom"] = whole_metrics.get("cases_ref_hom", MISSING_VALUE)
    row["cases_het"] = whole_metrics.get("cases_het", MISSING_VALUE)
    row["cases_alt_hom"] = whole_metrics.get("cases_alt_hom", MISSING_VALUE)
    row["cases_alt_freq"] = whole_metrics.get("cases_alt_freq", MISSING_VALUE)
    row["controls_ref_hom"] = whole_metrics.get("controls_ref_hom", MISSING_VALUE)
    row["controls_het"] = whole_metrics.get("controls_het", MISSING_VALUE)
    row["controls_alt_hom"] = whole_metrics.get("controls_alt_hom", MISSING_VALUE)
    row["controls_alt_freq"] = whole_metrics.get("controls_alt_freq", MISSING_VALUE)
    row["OR_whole_cohort"] = whole_metrics.get("OR_whole_cohort", MISSING_VALUE)
    row["warning"] = note or ""

    for label in ancestry_order:
        prefix = "{}_".format(label)
        values = metrics_by_ancestry[label]
        row[prefix + "n_cases_with_haplotype"] = values["n_cases_with_haplotype"]
        row[prefix + "n_controls_with_haplotype"] = values["n_controls_with_haplotype"]
        row[prefix + "total_hap_cases"] = values["total_hap_cases"]
        row[prefix + "total_hap_controls"] = values["total_hap_controls"]
        row[prefix + "cases_alt_count"] = values["cases_alt_count"]
        row[prefix + "cases_ref_count"] = values["cases_ref_count"]
        row[prefix + "controls_alt_count"] = values["controls_alt_count"]
        row[prefix + "controls_ref_count"] = values["controls_ref_count"]
        row[prefix + "cases_alt_freq"] = values["cases_alt_freq"]
        row[prefix + "controls_alt_freq"] = values["controls_alt_freq"]
        row[prefix + "OR"] = values["OR"]
        row[prefix + "dosage0_cases"] = values["dosage0_cases"]
        row[prefix + "dosage1_cases"] = values["dosage1_cases"]
        row[prefix + "dosage2_cases"] = values["dosage2_cases"]
        row[prefix + "dosage0_controls"] = values["dosage0_controls"]
        row[prefix + "dosage1_controls"] = values["dosage1_controls"]
        row[prefix + "dosage2_controls"] = values["dosage2_controls"]
        row[prefix + "warning"] = values["warning"]
        row[prefix + "tractor_p"] = per_ancestry_p.get(label, MISSING_VALUE)
    return row


def summary_dataframe_from_row(row, ancestry_order):
    columns = ["Whole cohort"] + ancestry_order
    data = OrderedDict()
    data["alt freq cases"] = [row.get("cases_alt_freq", MISSING_VALUE)] + [row.get("{}_cases_alt_freq".format(label), MISSING_VALUE) for label in ancestry_order]
    data["alt freq controls"] = [row.get("controls_alt_freq", MISSING_VALUE)] + [row.get("{}_controls_alt_freq".format(label), MISSING_VALUE) for label in ancestry_order]
    data["OR"] = [row.get("OR_whole_cohort", MISSING_VALUE)] + [row.get("{}_OR".format(label), MISSING_VALUE) for label in ancestry_order]
    data["Tractor p-value"] = [MISSING_VALUE] + [row.get("{}_tractor_p".format(label), MISSING_VALUE) for label in ancestry_order]
    data["N haplotypes cases"] = [2 * safe_float(row.get("n_cases_total", MISSING_VALUE)) if not pd.isna(row.get("n_cases_total", MISSING_VALUE)) else MISSING_VALUE] + [row.get("{}_total_hap_cases".format(label), MISSING_VALUE) for label in ancestry_order]
    data["N haplotypes controls"] = [2 * safe_float(row.get("n_controls_total", MISSING_VALUE)) if not pd.isna(row.get("n_controls_total", MISSING_VALUE)) else MISSING_VALUE] + [row.get("{}_total_hap_controls".format(label), MISSING_VALUE) for label in ancestry_order]
    frame = pd.DataFrame(data, index=columns).T
    return frame


def style_summary_sheet(ws, summary_df):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = BASE_FONT

    best_p = None
    best_cell = None
    for row_idx in range(2, ws.max_row + 1):
        metric = ws.cell(row=row_idx, column=1).value
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if metric == "Tractor p-value" and cell.value is not None:
                try:
                    numeric = float(cell.value)
                except Exception:
                    continue
                if best_p is None or numeric < best_p:
                    best_p = numeric
                    best_cell = cell
            if metric == "OR" and cell.value is not None:
                try:
                    numeric = float(cell.value)
                except Exception:
                    continue
                if numeric > 1.5:
                    cell.font = HIGHLIGHT_FONT
                    cell.fill = OR_HIGH_FILL
                elif numeric < 0.67:
                    cell.font = HIGHLIGHT_FONT
                    cell.fill = OR_LOW_FILL
    if best_cell is not None:
        best_cell.font = HIGHLIGHT_FONT
        best_cell.fill = P_BEST_FILL


def write_summary_card(row, ancestry_order, output_path, locus_name):
    summary_df = summary_dataframe_from_row(row, ancestry_order)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=sanitize_name(locus_name))
    workbook = load_workbook(output_path)
    worksheet = workbook[workbook.sheetnames[0]]
    style_summary_sheet(worksheet, summary_df)
    workbook.save(output_path)


def write_multi_sheet_workbook(rows, ancestry_order, output_path):
    workbook = Workbook()
    default_sheet = workbook.active
    if not rows:
        default_sheet.title = "summary"
        default_sheet.append(["message"])
        default_sheet.append(["No per-locus allele-count tables were available."])
        workbook.save(output_path)
        return
    workbook.remove(default_sheet)
    for row in rows:
        locus = str(row.get("locus", "locus"))
        sheet = workbook.create_sheet(title=sanitize_name(locus)[:31])
        summary_df = summary_dataframe_from_row(row, ancestry_order)
        sheet.append(["metric"] + list(summary_df.columns))
        for metric, values in summary_df.iterrows():
            sheet.append([metric] + list(values.values))
        style_summary_sheet(sheet, summary_df)
    workbook.save(output_path)


def locus_output_paths(outdir, locus):
    prefix = os.path.join(outdir, locus)
    return {
        "tract_plot": prefix + "_tract_plot.png",
        "composition_plot": prefix + "_ancestry_composition.png",
        "allele_freq_plot": prefix + "_allele_freq_by_ancestry.png",
        "counts_tsv": prefix + "_allele_counts.tsv",
        "summary_xlsx": prefix + "_summary_card.xlsx",
    }


def run_single_locus(args, runtime_cfg):
    os.makedirs(args.outdir, exist_ok=True)
    regions, norm_cols = load_regions(args.regions)
    region = get_region_row(regions, norm_cols, args.locus)
    fam_df, case_ids, control_ids = load_fam_status(runtime_cfg["fam_path"], runtime_cfg["case_code"], runtime_cfg["control_code"])
    n_case_rows = int((fam_df["PHENO"].astype(str) == str(runtime_cfg["case_code"])).sum())
    n_control_rows = int((fam_df["PHENO"].astype(str) == str(runtime_cfg["control_code"])).sum())
    lead, _ = find_lead_snp(region, runtime_cfg["ancestry_order"], runtime_cfg["ancestry_gwas_dir"])
    outputs = locus_output_paths(args.outdir, region["locus"])

    metrics_by_ancestry = OrderedDict()
    ancestry_rows = OrderedDict()
    per_ancestry_p = OrderedDict()
    note = ""

    if lead is None:
        warn("No lead SNP was found for locus {}; writing warning outputs".format(region["locus"]))
        note = "lead_snp_not_found"
        whole_metrics = {
            "n_cases_total": n_case_rows,
            "n_controls_total": n_control_rows,
            "cases_ref_hom": MISSING_VALUE,
            "cases_het": MISSING_VALUE,
            "cases_alt_hom": MISSING_VALUE,
            "cases_alt_freq": MISSING_VALUE,
            "controls_ref_hom": MISSING_VALUE,
            "controls_het": MISSING_VALUE,
            "controls_alt_hom": MISSING_VALUE,
            "controls_alt_freq": MISSING_VALUE,
            "OR_whole_cohort": MISSING_VALUE,
        }
        for label in runtime_cfg["ancestry_order"]:
            metrics_by_ancestry[label] = compute_ancestry_metrics(label, None, None, case_ids, control_ids)
            ancestry_rows[label] = {"dosage_row": None, "hapcount_row": None}
            per_ancestry_p[label] = MISSING_VALUE
        row = build_output_row(region, None, whole_metrics, runtime_cfg["ancestry_order"], metrics_by_ancestry, per_ancestry_p, note)
        pd.DataFrame([row]).to_csv(outputs["counts_tsv"], sep="\t", index=False)
        write_summary_card(row, runtime_cfg["ancestry_order"], outputs["summary_xlsx"], region["locus"])
        save_message_plot(outputs["tract_plot"], region["locus"], "Lead SNP was not found in any ancestry GWAS file.")
        save_message_plot(outputs["composition_plot"], region["locus"], "Lead SNP was not found in any ancestry GWAS file.")
        save_message_plot(outputs["allele_freq_plot"], region["locus"], "Lead SNP was not found in any ancestry GWAS file.")
        return

    for code in sorted(runtime_cfg["code_to_label"], key=lambda value: int(value)):
        label = runtime_cfg["code_to_label"][code]
        dosage_path = os.path.join(runtime_cfg["tractor_output_dir"], "chr{}.phased.anc{}.dosage.txt".format(region["chr"], code))
        hapcount_path = os.path.join(runtime_cfg["tractor_output_dir"], "chr{}.phased.anc{}.hapcount.txt".format(region["chr"], code))
        dosage_row = find_variant_row(dosage_path, lead["id"], lead["pos"])
        hapcount_row = find_variant_row(hapcount_path, lead["id"], lead["pos"])
        ancestry_rows[label] = {"dosage_row": dosage_row, "hapcount_row": hapcount_row}
        metrics_by_ancestry[label] = compute_ancestry_metrics(label, dosage_row, hapcount_row, case_ids, control_ids)
        gwas_row = find_gwas_value_for_lead(resolve_gwas_path(runtime_cfg["ancestry_gwas_dir"], label), lead["id"], lead["pos"])
        per_ancestry_p[label] = safe_float(gwas_row["_P"]) if gwas_row is not None else MISSING_VALUE
        if dosage_row is not None:
            if not lead.get("ref") and dosage_row.get("ref"):
                lead["ref"] = dosage_row.get("ref")
            if not lead.get("alt") and dosage_row.get("alt"):
                lead["alt"] = dosage_row.get("alt")

    vcf_path = os.path.join(runtime_cfg["step4_output_dir"], "chr{}.phased.vcf.gz".format(region["chr"]))
    whole_metrics = compute_whole_cohort_from_vcf(vcf_path, region["chr"], lead["id"], lead["pos"], case_ids, control_ids)
    if whole_metrics is None:
        whole_metrics = compute_whole_cohort_from_reconstructed(ancestry_rows, case_ids, control_ids) or {}
    if not lead.get("ref") and whole_metrics.get("ref"):
        lead["ref"] = whole_metrics.get("ref")
    if not lead.get("alt") and whole_metrics.get("alt"):
        lead["alt"] = whole_metrics.get("alt")

    row = build_output_row(region, lead, whole_metrics, runtime_cfg["ancestry_order"], metrics_by_ancestry, per_ancestry_p, note)
    pd.DataFrame([row]).to_csv(outputs["counts_tsv"], sep="\t", index=False)
    write_summary_card(row, runtime_cfg["ancestry_order"], outputs["summary_xlsx"], region["locus"])

    msp_path = os.path.join(runtime_cfg["step5_output_dir"], "chr{}.deconvoluted.msp.tsv".format(region["chr"]))
    make_tract_plot(
        outputs["tract_plot"],
        region,
        lead["id"],
        lead["pos"],
        runtime_cfg["ancestry_order"],
        runtime_cfg["code_to_label"],
        msp_path,
        case_ids,
        control_ids,
        args.n_samples,
    )
    make_ancestry_composition_plot(outputs["composition_plot"], region, runtime_cfg["ancestry_order"], metrics_by_ancestry)
    make_allele_frequency_plot(outputs["allele_freq_plot"], region, runtime_cfg["ancestry_order"], metrics_by_ancestry, whole_metrics)


def discover_count_files(counts_dir, loci_order):
    discovered = {}
    for locus in loci_order:
        path = os.path.join(counts_dir, str(locus), "{}_allele_counts.tsv".format(locus))
        if os.path.exists(path):
            discovered[str(locus)] = path
    if not discovered:
        for root, _, files in os.walk(counts_dir):
            for name in sorted(files):
                if name.endswith("_allele_counts.tsv"):
                    locus = name[: -len("_allele_counts.tsv")]
                    discovered[locus] = os.path.join(root, name)
    return discovered


def run_merge_only(args, runtime_cfg):
    os.makedirs(args.outdir, exist_ok=True)
    regions, norm_cols = load_regions(args.regions)
    loci_order = [str(value) for value in regions[norm_cols["LOCUS"]].dropna().tolist()]
    count_files = discover_count_files(args.counts_dir, loci_order)
    frames = []
    ordered_rows = []
    for locus in loci_order:
        path = count_files.get(locus)
        if not path:
            warn("Missing allele-count table for locus {} in {}".format(locus, args.counts_dir))
            continue
        frame = pd.read_csv(path, sep="\t")
        if frame.empty:
            continue
        frames.append(frame)
        ordered_rows.append(frame.iloc[0].to_dict())

    merged_tsv = os.path.join(args.outdir, "all_loci_allele_counts.tsv")
    merged_xlsx = os.path.join(args.outdir, "all_loci_summary_cards.xlsx")
    if frames:
        merged = pd.concat(frames, ignore_index=True)
    else:
        merged = pd.DataFrame()
    merged.to_csv(merged_tsv, sep="\t", index=False)
    write_multi_sheet_workbook(ordered_rows, runtime_cfg["ancestry_order"], merged_xlsx)


def main():
    args = parse_args()
    config_path = resolve_config_path(args.config)
    cfg = load_yaml(config_path)
    runtime_cfg = derive_runtime_config(cfg, args.regions)
    if args.merge_only:
        run_merge_only(args, runtime_cfg)
    else:
        run_single_locus(args, runtime_cfg)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("locus_haplotype_microscope failed: {}".format(exc), file=sys.stderr)
        sys.exit(1)